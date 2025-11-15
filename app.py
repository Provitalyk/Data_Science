import pandas as pd
import numpy as np
import warnings

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
warnings.filterwarnings('ignore')

# Загрузка данных
try:
    prolongations = pd.read_csv('prolongations.csv')
    financial_data = pd.read_csv('financial_data.csv')
except FileNotFoundError as e:
    print("❌ Ошибка: не найдены файлы prolongations.csv или financial_data.csv")
    exit()

print("✅ Данные загружены")
print("prolongations.head():")
print(prolongations.head())
print("\nfinancial_data.head():")
print(financial_data.head())

# Очистка названий колонок
prolongations.columns = prolongations.columns.str.strip()
financial_data.columns = financial_data.columns.str.strip()

# Приведение id к числовому типу
prolongations['id'] = pd.to_numeric(prolongations['id'], errors='coerce')
financial_data['id'] = pd.to_numeric(financial_data['id'], errors='coerce')

# Автоопределение колонки AM
am_col = None
for col in prolongations.columns:
    if any(k in col.lower() for k in ['ам', 'am', 'менеджер', 'account']):
        am_col = col
        break
if not am_col:
    raise ValueError(f"❌ Не найдена колонка с аккаунт-менеджером. Доступные: {prolongations.columns.tolist()}")
prolongations = prolongations.rename(columns={am_col: 'AM'})
print(f"✅ Колонка '{am_col}' переименована в 'AM'")

# Поиск колонок-месяцев
months_map = {
    'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6,
    'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12
}

month_cols = []
col_to_date = {}
for col in financial_data.columns:
    clean = col.strip()
    if ' ' not in clean:
        continue
    parts = clean.split()
    if len(parts) != 2:
        continue
    m_name, y_str = parts[0].lower(), parts[1]
    if m_name in months_map and y_str.isdigit():
        month_cols.append(col)
        col_to_date[col] = (int(y_str), months_map[m_name])

if not month_cols:
    raise ValueError("❌ Не найдены колонки с месяцами в формате 'Месяц Год'")

print(f"✅ Найдено {len(month_cols)} колонок с месяцами")
for col, (y, m) in col_to_date.items():
    print(f"   {col} → {y}-{m:02d}")

# Перевод в long-формат
financial_long = financial_data[['id', 'Account', 'Причина дубля'] + month_cols] \
    .melt(id_vars=['id', 'Account', 'Причина дубля'],
          value_vars=month_cols,
          var_name='month_col',
          value_name='shipment')

financial_long['year'] = financial_long['month_col'].map(lambda x: col_to_date[x][0])
financial_long['month'] = financial_long['month_col'].map(lambda x: col_to_date[x][1])
financial_long = financial_long.sort_values(['id', 'year', 'month']).reset_index(drop=True)

# Обработка значений
def parse_shipment(val):
    if pd.isna(val):
        return np.nan
    if isinstance(val, str):
        val = val.strip().lower()
        if val in ['стоп', 'end']:
            return 'stop'
        elif val == 'в ноль':
            return 0.0
        else:
            try:
                return float(val)
            except:
                return np.nan
    return val

financial_long['parsed'] = financial_long['shipment'].apply(parse_shipment)
financial_long['is_stop'] = financial_long['parsed'] == 'stop'
# Заменяем 'stop' на NaN, остальное — в числа (включая 0)
financial_long['value'] = pd.to_numeric(
    financial_long['parsed'].replace('stop', np.nan),
    errors='coerce'
)

# Определение последнего активного месяца реализации (без учёта 'стоп')
active = financial_long.dropna(subset=['value'])  # сохраняем 0, убираем NaN
last_active = active.groupby('id').apply(
    lambda g: g.loc[g['year'] * 100 + g['month'] == (g['year'] * 100 + g['month']).max()]
).reset_index(drop=True)[['id', 'year', 'month']].rename(columns={'year': 'final_year', 'month': 'final_month'})

# Обработка prolongations: извлечение даты из текста
def parse_prolong_month(text):
    if pd.isna(text):
        return None, None
    parts = str(text).strip().split()
    if len(parts) != 2:
        return None, None
    m_name, y_str = parts[0].lower(), parts[1]
    if m_name in months_map and y_str.isdigit():
        return int(y_str), months_map[m_name]
    return None, None

prolongations[['p_year', 'p_month']] = prolongations['month'].apply(
    lambda x: pd.Series(parse_prolong_month(x))
)

# Присоединение real final date
prolongations = prolongations.merge(last_active, on='id', how='left')

# Исключение проектов с 'стоп' в последний месяц или ранее
# Ищем последний период реализации
project_final_period = active.groupby('id').apply(lambda g: g['year'].iloc[-1] * 100 + g['month'].iloc[-1])

# Найдём первый 'стоп'
stop_log = financial_long[financial_long['is_stop']].copy()
stop_log['period'] = stop_log['year'] * 100 + stop_log['month']
stop_min = stop_log.groupby('id')['period'].min()

# Фильтр: только если 'стоп' был после последнего месяца
valid_ids = [
    pid for pid in project_final_period.index
    if pid not in stop_min or stop_min[pid] > project_final_period[pid]
]

financial_long = financial_long[financial_long['id'].isin(valid_ids)]

# Добавление AM и менеджера
financial_long = financial_long.merge(prolongations[['id', 'AM']], on='id', how='left')
financial_long['manager'] = financial_long['AM'].fillna(financial_long['Account'])

# Добавление final_year и final_month
final_dates = prolongations[['id', 'final_year', 'final_month']].drop_duplicates()
financial_long = financial_long.merge(final_dates, on='id', how='left')

# Анализ пролонгаций
def add_months(year, month, add):
    total = year * 12 + month - 1 + add
    return total // 12, (total % 12) + 1


def analyze_project(group):
    if group['final_year'].isna().any() or group['final_month'].isna().any():
        return pd.Series({'ship_last': 0, 'ship_m1': 0, 'ship_m2': 0})
    try:
        fy, fm = int(group['final_year'].iloc[0]), int(group['final_month'].iloc[0])
    except:
        return pd.Series({'ship_last': 0, 'ship_m1': 0, 'ship_m2': 0})

    # Проверим: если в final_month у всех строк shipment = 'в ноль' или 0
    final_month_data = group[(group['year'] == fy) & (group['month'] == fm)]

    # Если все строки в final_month — 'в ноль' или 0
    if len(final_month_data) > 0:
        all_zero = True
        for val in final_month_data['parsed']:
            if val not in [0, 'в ноль']:
                all_zero = False
                break
        if all_zero and len(final_month_data) > 0:
            # Берём отгрузку из предыдущего месяца
            prev_y, prev_m = add_months(fy, fm, -1)
            prev_data = group[(group['year'] == prev_y) & (group['month'] == prev_m)]
            ship_last = prev_data['value'].sum()
        else:
            ship_last = final_month_data['value'].sum()
    else:
        ship_last = 0

    y1, m1 = add_months(fy, fm, 1)
    y2, m2 = add_months(fy, fm, 2)

    ship_m1 = group[(group['year'] == y1) & (group['month'] == m1)]['value'].sum()
    ship_m2 = group[(group['year'] == y2) & (group['month'] == m2)]['value'].sum()

    return pd.Series({
        'final_year': fy,
        'final_month': fm,
        'ship_last': ship_last,
        'ship_m1': ship_m1,
        'ship_m2': ship_m2
    })

# Группируем
stats = financial_long.groupby(['id', 'manager']).apply(analyze_project).reset_index()

# Фильтр: только проекты, где M+1 или M+2 — в 2023 году
valid_stats = []
for _, row in stats.iterrows():
    if pd.isna(row['final_year']) or pd.isna(row['final_month']):
        continue
    fy, fm = int(row['final_year']), int(row['final_month'])
    y1, m1 = add_months(fy, fm, 1)
    y2, m2 = add_months(fy, fm, 2)
    if y1 == 2023 or y2 == 2023:
        row['final_year'] = fy
        row['final_month'] = fm
        valid_stats.append(row)

if not valid_stats:
    print("❌ Нет проектов с пролонгацией в 2023 году.")
    exit()

stats = pd.DataFrame(valid_stats)
stats['final_year'] = stats['final_year'].astype(int)
stats['final_month'] = stats['final_month'].astype(int)

print(f"✅ Найдено {len(stats)} проектов с пролонгацией в 2023 году")

# Диагностика
print("\n" + "="*60)
print("🔍 ДИАГНОСТИКА: ПРИМЕРЫ ПРОЕКТОВ")
print("="*60)
for _, row in stats.head(5).iterrows():
    print(f"ID: {row['id']}, Менеджер: {row['manager']}")
    print(f"  Завершён: {row['final_year']}-{row['final_month']:02d}")
    print(f"  Отгрузка в последний месяц: {row['ship_last']:,.0f}")
    print(f"  Отгрузка в M+1: {row['ship_m1']:,.0f} → {'Да' if row['ship_m1'] > 0 else 'Нет'}")
    print(f"  Отгрузка в M+2: {row['ship_m2']:,.0f} → {'Да' if row['ship_m2'] > 0 else 'Нет'}")
    print("-" * 40)

# Месячные коэффициенты (февраль — декабрь 2023)
monthly = []
for m in range(2, 13):  # анализ за фев — дек 2023
    prev_y, prev_m = add_months(2023, m, -1)
    df_m1 = stats[(stats['final_year'] == prev_y) & (stats['final_month'] == prev_m)]
    base_m1 = df_m1['ship_last'].sum()
    ship_m1 = df_m1['ship_m1'].sum()
    coef_m1 = ship_m1 / base_m1 if base_m1 > 0 else 0.0

    prev2_y, prev2_m = add_months(2023, m, -2)
    df_m2_all = stats[(stats['final_year'] == prev2_y) & (stats['final_month'] == prev2_m)]
    df_m2_no = df_m2_all[df_m2_all['ship_m1'] == 0]
    base_m2 = df_m2_no['ship_last'].sum()
    ship_m2 = df_m2_no['ship_m2'].sum()
    coef_m2 = ship_m2 / base_m2 if base_m2 > 0 else 0.0

    monthly.append({
        'Месяц': ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'][m-1],
        'Коэфф. M1': round(coef_m1, 3),
        'Коэфф. M2': round(coef_m2, 3),
        'База (M-1)': int(base_m1),
        'Продлено (M+1)': int(ship_m1),
        'База (M-2, без M+1)': int(base_m2),
        'Продлено (M+2)': int(ship_m2),
    })

monthly_df = pd.DataFrame(monthly)

# Годовые коэффициенты по менеджерам
annual = []
for mgr in stats['manager'].dropna().unique():
    data = stats[stats['manager'] == mgr]
    # M1: финал в янв–ноя
    d1 = data[data['final_month'].between(1, 11)]
    base1 = d1['ship_last'].sum()
    ship1 = d1['ship_m1'].sum()
    coef1 = ship1 / base1 if base1 > 0 else 0.0
    # M2: финал в янв–окт, без M+1
    d2 = data[data['final_month'].between(1, 10)]
    d2_no = d2[d2['ship_m1'] == 0]
    base2 = d2_no['ship_last'].sum()
    ship2 = d2_no['ship_m2'].sum()
    coef2 = ship2 / base2 if base2 > 0 else 0.0

    annual.append({
        'Менеджер': mgr,
        'Коэфф. M1 (год)': round(coef1, 3),
        'Коэфф. M2 (год)': round(coef2, 3),
        'База M1': int(base1),
        'Продлено M1': int(ship1),
        'База M2': int(base2),
        'Продлено M2': int(ship2),
    })

annual_df = pd.DataFrame(annual)

# Весь отдел
total_base1 = annual_df['База M1'].sum()
total_ship1 = annual_df['Продлено M1'].sum()
total_base2 = annual_df['База M2'].sum()
total_ship2 = annual_df['Продлено M2'].sum()
coef_m1_total = total_ship1 / total_base1 if total_base1 > 0 else 0.0
coef_m2_total = total_ship2 / total_base2 if total_base2 > 0 else 0.0

annual_df.loc[len(annual_df)] = {
    'Менеджер': 'Весь отдел',
    'Коэфф. M1 (год)': round(coef_m1_total, 3),
    'Коэфф. M2 (год)': round(coef_m2_total, 3),
    'База M1': int(total_base1),
    'Продлено M1': int(total_ship1),
    'База M2': int(total_base2),
    'Продлено M2': int(total_ship2),
}

# Создаём Excel-файл
wb = Workbook()

# === Лист 1: Обзор ===
ws_overview = wb.active
ws_overview.title = "Обзор"

# Заголовок
ws_overview['A1'] = "Отчёт по пролонгации договоров — 2023 год"
ws_overview['A1'].font = Font(size=16, bold=True)
ws_overview['A2'] = f"Подготовлено: {pd.Timestamp.now().strftime('%d.%m.%Y')}"
ws_overview['A4'] = "Ключевые метрики"

# Стиль для заголовков
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")

# Ключевые показатели
overview_data = [
    ["Показатель", "Значение"],
    ["Коэффициент M1 (средний)", round(annual_df[annual_df['Менеджер'] == 'Весь отдел']['Коэфф. M1 (год)'].iloc[0], 3)],
    ["Коэфф. M2 (средний)", round(annual_df[annual_df['Менеджер'] == 'Весь отдел']['Коэфф. M2 (год)'].iloc[0], 3)],
    ["Всего проектов", len(stats)],
    ["Успешно пролонгировано в M+1", f"{round(annual_df[annual_df['Менеджер'] == 'Весь отдел']['Коэфф. M1 (год)'].iloc[0] * 100, 1)}%"],
    ["Успешно в M+2", f"{round(annual_df[annual_df['Менеджер'] == 'Весь отдел']['Коэфф. M2 (год)'].iloc[0] * 100, 1)}%"]
]

for r_idx, row in enumerate(overview_data, 5):
    for c_idx, value in enumerate(row, 1):
        cell = ws_overview.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 5:  # Заголовки
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.font = Font(size=11)

# Расширяем столбцы
ws_overview.column_dimensions['A'].width = 30
ws_overview.column_dimensions['B'].width = 15

# === Лист 2: Месячные ===
ws_month = wb.create_sheet("Месячные")
for r in dataframe_to_rows(monthly_df, index=False, header=True):
    ws_month.append(r)

# Стиль заголовков
for cell in ws_month[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Форматирование чисел
for row in ws_month.iter_rows(min_row=2, min_col=3, max_col=7):
    for cell in row:
        cell.number_format = '#,##0'

# Цвета
m1_col = ws_month.column_dimensions['B']
m2_col = ws_month.column_dimensions['C']
m1_col.font = Font(color="006400")
m2_col.font = Font(color="000080")

# === Лист 3: Годовые ===
ws_annual = wb.create_sheet("Годовые")
for r in dataframe_to_rows(annual_df, index=False, header=True):
    ws_annual.append(r)

# Стиль заголовков
for cell in ws_annual[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Форматирование
for row in ws_annual.iter_rows(min_row=2, min_col=3, max_col=8):
    for cell in row:
        cell.number_format = '#,##0'

# Выделение "Весь отдел"
for cell in ws_annual[ws_annual.max_row]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

# === Графики ===

# График M1 и M2 по месяцам
chart1 = LineChart()
chart1.title = "Коэффициенты пролонгации по месяцам"
chart1.style = 12
chart1.y_axis.title = 'Коэффициент'
chart1.x_axis.title = 'Месяц'

cats = Reference(ws_month, min_col=1, min_row=2, max_row=len(monthly_df)+1)
data = Reference(ws_month, min_col=2, max_col=3, min_row=1, max_row=len(monthly_df)+1)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.width = 20
chart1.height = 8
ws_overview.add_chart(chart1, "E1")

# График: Топ-5 менеджеров по M1
top5 = annual_df[annual_df['Менеджер'] != 'Весь отдел'].sort_values('Коэфф. M1 (год)', ascending=False).head(5)
for idx, row in top5.iterrows():
    ws_annual.append([row['Менеджер'], row['Коэфф. M1 (год)'], row['Коэфф. M2 (год)']])

chart2 = BarChart()
chart2.type = "col"
chart2.style = 6
chart2.title = "Топ-5 менеджеров по M1"
chart2.y_axis.title = 'Коэффициент'

cats2 = Reference(ws_annual, min_col=1, min_row=ws_annual.max_row-4, max_row=ws_annual.max_row)
data2 = Reference(ws_annual, min_col=2, min_row=ws_annual.max_row-5, max_row=ws_annual.max_row)
chart2.add_data(data2, titles_from_data=False)
chart2.set_categories(cats2)
chart2.shape = 4
ws_overview.add_chart(chart2, "E20")

# Сохранение
wb.save('отчет_пролонгации_2023.xlsx')

print("✅ Отчёт успешно сохранён: отчет_пролонгации_2023.xlsx")